import openpyxl
import config
from openpyxl.styles import colors, Font, PatternFill
import os
import time


class Execl_open:

    def __init__(self, filename=os.path.join(config.testdatas_dir, "testcase.xlsx")):
        self.filepath = filename
        self.wk = openpyxl.load_workbook(filename)

    def get_data(self):
        """获取execl工作簿中要执行用例的数据"""
        casesname = self.get_casename()
        self.sheetnames = casesname
        values = []
        for sheetname in self.sheetnames:
            value = self.get_startcol_endcol_value(sheetname)
            # 将表名和表数据组成一个字典
            casedata = {}
            casedata['casename'] = sheetname
            casedata['stepdata'] = value
            values.append(casedata)
        return values

    # 获取部分区域表的全部值
    def get_startcol_endcol_value(self, sheetname, start=config.keyword_col, end=config.opera_col):

        sheet = self.wk[sheetname]
        values = []
        for row in range(2, sheet.max_row + 1):
            value_data = []
            for col in range(start, end + 1):
                value = sheet.cell(row, col).value
                if value != None:
                    value_data.append(value)
            values.append(value_data)
        return values

    # 获取执行用例表名
    def get_casename(self, sheetname=config.casesheetsname):
        sheet = self.wk[sheetname]
        values = []
        for row in range(2, sheet.max_row + 1):
            isexecute_col = sheet.cell(row, config.isexecut_col).value
            if isexecute_col == 'y':
                value = sheet.cell(row, config.step_col).value
                values.append(value)
        return values

    # 写入执行结果
    def write_stepresult(self, sheetname, row, col, result):
        casesheet=self.wk[sheetname]
        casesheet.cell(row, col).value = result
        # 红色填充到单元格
        readfill = PatternFill("solid", fgColor="00FF0000")
        # 绿色填充到单元格
        greenfill = PatternFill("solid", fgColor="0000FF00")
        if result == 'FLASE':
            casesheet.cell(row, col).fill = readfill
        else:
            casesheet.cell(row, col).fill = greenfill
        self.wk.save(self.filepath)


    # 写入汇总表执行结果
    def write_caseresult(self, sheetname=config.casesheetsname, col=config.caseresult_col):

        sheet = self.wk[sheetname]
        for row in range(2, sheet.max_row + 1):
            # 红色填充到单元格
            readfill = PatternFill("solid", fgColor="00FF0000")
            # 绿色填充到单元格
            greenfill = PatternFill("solid", fgColor="0000FF00")
            # 是否执行的值
            isresult = sheet.cell(row, config.isexecut_col).value
            if isresult == 'y':
                casename = sheet.cell(row, config.step_col).value
                result = self.get_stepresult(casename)
                if 'FLASE' in result:
                    sheet.cell(row, col).fill = readfill
                    sheet.cell(row, col).value = 'FLASE'
                else:
                    sheet.cell(row, col).fill = greenfill
                    sheet.cell(row, col).value = 'PASS'
        #self.wk.save(self.filepath)

    # 获取单个用例shepresult结果
    def get_stepresult(self, sheetname):
        sheet = self.wk[sheetname]
        values = []
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, config.stepresult_col).value
            values.append(value)
        return values

    # 写入执行时间

    def write_steptime(self, sheetname, row, col=config.steptime_col):
        sheet = self.wk[sheetname]
        step_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        sheet.cell(row, col).value = step_time
        self.wk.save(self.filepath)


    # 写入汇总表执行时间
    def write_casetime(self, sheetname=config.casesheetsname,col=config.casetime_col):

        sheet = self.wk[sheetname]
        step_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        for row in range(2, sheet.max_row + 1):

            # 是否执行的值
            isresult = sheet.cell(row, config.isexecut_col).value
            if isresult == 'y':
                sheet.cell(row,col).value = step_time

    #


if __name__ == '__main__':
    print(Execl_open().get_data())
